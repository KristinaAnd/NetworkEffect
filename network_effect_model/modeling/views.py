from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Avg
from django.http import HttpResponse
from .models import GraphModel
from .services.graph_generator import GraphGenerator
from .services.supply_chain_analyzer import SupplyChainAnalyzer
import networkx as nx
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import base64
from io import BytesIO
import json
import csv
from datetime import datetime
from scipy import stats
import numpy as np

# ---------- ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ----------
def visualize_graph(G, directed, weighted, balanced=False):
    """Визуализация графа с отображением весов на рёбрах (всегда)."""
    plt.figure(figsize=(12, 9))
    pos = nx.spring_layout(G, k=1.5, iterations=40, seed=42)
    nx.draw_networkx_nodes(G, pos, node_color='#3498db', node_size=400, alpha=0.9,
                           edgecolors='#2c3e50', linewidths=1.5)
    nx.draw_networkx_labels(G, pos, font_size=10, font_weight='bold')

    # Определяем ширину линий в зависимости от веса
    if weighted:
        edge_widths = [min(data.get('weight', 0.5) * 0.8, 1.2) for _, _, data in G.edges(data=True)]
        edge_colors = ['#7f8c8d' for _ in G.edges()]
    else:
        edge_widths = [0.6 for _ in G.edges()]
        edge_colors = '#7f8c8d'

    if directed:
        if balanced:
            # Двусторонние стрелки для симметричного направленного графа
            for u, v, data in G.edges(data=True):
                width = data.get('weight', 1) * 0.8 if weighted else 0.6
                nx.draw_networkx_edges(G, pos, edgelist=[(u, v)], arrows=True, arrowstyle='->',
                                       arrowsize=10, width=width, alpha=0.6, edge_color='#7f8c8d',
                                       connectionstyle="arc3,rad=0.1")
                if G.has_edge(v, u) or balanced:
                    nx.draw_networkx_edges(G, pos, edgelist=[(v, u)], arrows=True, arrowstyle='->',
                                           arrowsize=10, width=width, alpha=0.6, edge_color='#7f8c8d',
                                           connectionstyle="arc3,rad=-0.1")
        else:
            nx.draw_networkx_edges(G, pos, arrowsize=10, width=edge_widths, alpha=0.6,
                                   edge_color=edge_colors, arrowstyle='->', connectionstyle='arc3,rad=0.1')
    else:
        nx.draw_networkx_edges(G, pos, width=edge_widths, alpha=0.6, edge_color=edge_colors)

    # Отображаем веса на рёбрах ВСЕГДА (для взвешенных – реальные, для невзвешенных – 1)
    for u, v, data in G.edges(data=True):
        w = data.get('weight', 1)
        mid_x = (pos[u][0] + pos[v][0]) / 2
        mid_y = (pos[u][1] + pos[v][1]) / 2
        plt.text(mid_x, mid_y, f"{w:.2f}", fontsize=8, ha='center', va='center',
                 bbox=dict(boxstyle='round,pad=0.2', facecolor='white', alpha=0.8))

    title = f"{'Направленный' if directed else 'Ненаправленный'} граф"
    if weighted:
        title += " взвешенный "
    if balanced and directed:
        title += "и симметричный"
    plt.title(title, fontsize=12, fontweight='bold', pad=15)
    plt.axis('off')
    buffer = BytesIO()
    plt.savefig(buffer, format='png', dpi=150, bbox_inches='tight', facecolor='white')
    buffer.seek(0)
    image_base64 = base64.b64encode(buffer.getvalue()).decode()
    plt.close()
    return image_base64


# ---------- ОСНОВНЫЕ ВИДЫ ----------
def dashboard(request):
    total_models = GraphModel.objects.count()
    avg_effect = GraphModel.objects.filter(total_effect__isnull=False).aggregate(Avg('total_effect'))['total_effect__avg'] or 0
    avg_structural = GraphModel.objects.filter(structural_effect__isnull=False).aggregate(Avg('structural_effect'))['structural_effect__avg'] or 0
    avg_functional = GraphModel.objects.filter(functional_effect__isnull=False).aggregate(Avg('functional_effect'))['functional_effect__avg'] or 0

    models_data = []
    density_list = []
    effect_list = []
    agents_list = []
    weight_avg_list = []
    weight_effect_list = []

    for m in GraphModel.objects.all():
        if m.total_effect is not None:
            density = m.get_density()
            models_data.append({
                'id': m.id,
                'num_agents': m.num_agents,
                'num_edges': m.num_edges,
                'density': density,
                'directed': m.directed,
                'weight_type': m.weight_type,
                'weight_avg': m.weight_avg if m.weight_avg is not None else 1.0,
                'total_effect': m.total_effect,
                'structural_effect': m.structural_effect or 0,
                'functional_effect': m.functional_effect or 0,
            })
            density_list.append(density)
            effect_list.append(m.total_effect)
            agents_list.append(m.num_agents)
            if m.weight_type == 'weighted' and m.weight_avg:
                weight_avg_list.append(m.weight_avg)
                weight_effect_list.append(m.total_effect)

    # Корреляции и регрессии
    density_corr = np.corrcoef(density_list, effect_list)[0, 1] if len(density_list) > 1 else 0
    agents_corr = np.corrcoef(agents_list, effect_list)[0, 1] if len(agents_list) > 1 else 0
    weight_corr = np.corrcoef(weight_avg_list, weight_effect_list)[0, 1] if len(weight_avg_list) > 1 else 0

    if len(set(density_list)) > 1:
       density_corr = np.corrcoef(density_list, effect_list)[0, 1]
       slope, intercept, _, _, _ = stats.linregress(density_list, effect_list)
       density_slope, density_intercept = slope, intercept
    else:
       density_corr = 0.0
       density_slope = density_intercept = 0.0
    
    if len(set(agents_list)) > 1:
       agents_corr = np.corrcoef(agents_list, effect_list)[0, 1]
       slope, intercept, _, _, _ = stats.linregress(agents_list, effect_list)
       agents_slope, agents_intercept = slope, intercept
    else:
       agents_corr = 0.0
       agents_slope = agents_intercept = 0.0
    
    if len(set(weight_avg_list)) > 1:
       weight_corr = np.corrcoef(weight_avg_list, weight_effect_list)[0, 1]
       slope, intercept, _, _, _ = stats.linregress(weight_avg_list, weight_effect_list)
       weight_slope, weight_intercept = slope, intercept
    else:
       weight_corr = 0.0
       weight_slope = weight_intercept = 0.0

    effect_arr = np.array(effect_list) if effect_list else [0]
    median_effect = np.median(effect_arr)
    min_effect = np.min(effect_arr) if len(effect_arr) > 0 else 0
    max_effect = np.max(effect_arr) if len(effect_arr) > 0 else 0
    std_effect = np.std(effect_arr)
    percentile_25 = np.percentile(effect_arr, 25) if len(effect_arr) > 0 else 0
    percentile_75 = np.percentile(effect_arr, 75) if len(effect_arr) > 0 else 0
    growth_potential = max_effect - min_effect
    growth_percent = (growth_potential / max_effect * 100) if max_effect > 0 else 0

    context = {
        'total_models': total_models,
        'avg_effect': round(avg_effect, 4),
        'avg_structural': round(avg_structural, 4),
        'avg_functional': round(avg_functional, 4),
        'recent_models': GraphModel.objects.all()[:5],
        'models_json': json.dumps(models_data),
        'directed_count': GraphModel.objects.filter(directed='directed').count(),
        'weighted_count': GraphModel.objects.filter(weight_type='weighted').count(),
        'density_correlation': density_corr,
        'agents_correlation': agents_corr,
        'weight_correlation': weight_corr,
        'density_slope': density_slope,
        'density_intercept': density_intercept,
        'agents_slope': agents_slope,
        'agents_intercept': agents_intercept,
        'weight_slope': weight_slope,
        'weight_intercept': weight_intercept,
        'median_effect': median_effect,
        'min_effect': min_effect,
        'max_effect': max_effect,
        'std_effect': std_effect,
        'percentile_25': percentile_25,
        'percentile_75': percentile_75,
        'growth_potential': growth_potential,
        'growth_percent': growth_percent,
    }
    return render(request, 'modeling/dashboard.html', context)


def generate_form(request):
    return render(request, 'modeling/generate_form.html')


def generate_submit(request):
    if request.method != 'POST':
        return redirect('modeling:generate_form')

    try:
        num_agents = int(request.POST.get('num_agents', 10))
        if num_agents > 100:
            messages.error(request, 'Количество агентов не может превышать 100')
            return redirect('modeling:generate_form')
        num_edges = int(request.POST.get('density', 10))
        directed = request.POST.get('directed') == 'directed'
        weight_type = request.POST.get('weight_type', 'unweighted')
        balanced = request.POST.get('balance_type', 'balanced') == 'balanced'
        weight_min = float(request.POST.get('weight_min', 0.1))
        weight_max = float(request.POST.get('weight_max', 1.0))
        is_random = request.POST.get('random', 'false') == 'true'
        model_name = request.POST.get('name', f'Модель {datetime.now().strftime("%H:%M:%S")}')

        generator = GraphGenerator(num_agents, num_edges, directed,
                                   weight_type == 'weighted', balanced,
                                   weight_min, weight_max)
        if is_random:
            G = generator.generate_random()
        else:
            G = generator.generate()

        stats = generator.get_stats(G)
        structural, functional, total, B = generator.calculate_effects(G)

        # Получаем плотность D из stats
        D = stats['density_percent'] / 100.0

        # Анализ цепочки поставок (научно обоснованные метрики)
        analyzer = SupplyChainAnalyzer(G, directed, weight_type == 'weighted')
        analysis = analyzer.analyze_efficiency()
        efficiency_score = analysis['efficiency_score']
        optimization_suggestions = analysis['optimization_suggestions']

        # Бизнес-ценность по закону Меткалфа: V = TNE * N^2
        business_value = round(total * (num_agents ** 2), 4)

        # Уровень риска: R = (1 - B) * 100%   (B – сбалансированность)
        risk_score = round((1 - (D * B) ** 0.5) * 100, 1)
        risk_score = min(max(risk_score, 0), 100)

        # Синергия (мультипликатор) – линейная (консервативная оценка)
        synergy_factor = 1 + (num_agents - 2) * 0.05
        # Экономический эффект от добавления одного агента
        potential_growth_raw = ((num_agents + 1) ** 2 / (num_agents ** 2) - 1)
        if num_agents > 0:
            economic_impact = economic_impact = round(potential_growth_raw * efficiency_score * 100, 2)
        else:
            economic_impact = 0

        metcalfe_value = num_agents ** 2
        max_possible = num_agents * (num_agents - 1)
        if not directed:
            max_possible //= 2
        scale_efficiency = (num_edges / max_possible * 100) if max_possible > 0 else 0

        # Интерпретации
        interpretations = []
        if total >= 60:
            interpretations.append("Высокий сетевой эффект — система демонстрирует синергию")
        elif total >= 30:
            interpretations.append("Средний сетевой эффект — потенциал раскрыт не полностью")
        else:
            interpretations.append("Низкий сетевой эффект — синергия отсутствует")
        if B > 80:
            interpretations.append(f"Сбалансированная структура (B={B:.3f}) — равномерное распределение связей")
        elif B > 40:
            interpretations.append(f"Умеренно сбалансированная структура (B={B:.3f})")
        else:
            interpretations.append(f"Несбалансированная структура (B={B:.3f}) — высокая концентрация на хабах")
        interpretations.append(f"Закон Меткалфа: ценность сети ~ n² = {metcalfe_value}")

        recommendations_text = '\n'.join(optimization_suggestions) + '\n\n' + '\n'.join(interpretations)

        model = GraphModel.objects.create(
            name=model_name,
            num_agents=num_agents,
            num_edges=num_edges,
            directed='directed' if directed else 'undirected',
            weight_type=weight_type,
            balance_type='balanced' if balanced else 'unbalanced',
            weight_min=weight_min if weight_type == 'weighted' else None,
            weight_max=weight_max if weight_type == 'weighted' else None,
            weight_avg=stats.get('weight_avg', 1.0),
            structural_effect=structural,
            functional_effect=functional,
            total_effect=total,
            efficiency=efficiency_score,
            business_value=business_value,
            risk_score=risk_score,
            synergy_factor=synergy_factor,
            economic_impact=economic_impact,
            recommendation=recommendations_text[:2000],
        )

        # Сохраняем структуру графа
        edges_data = []
        for u, v, data in G.edges(data=True):
            edge = {'source': u, 'target': v}
            if weight_type == 'weighted' and 'weight' in data:
                edge['weight'] = data['weight']
            else:
                edge['weight'] = 1.0   # невзвешенный – вес 1
            edges_data.append(edge)
        model.graph_data = {'nodes': list(G.nodes()), 'edges': edges_data}
        model.save()

        image_base64 = visualize_graph(G, directed, weight_type == 'weighted', balanced)

        # Градация
        if total >= 60:
            grade = "Высокий"
            grade_color = "#27ae60"
        elif total >= 30:
            grade = "Средний"
            grade_color = "#f39c12"
        else:
            grade = "Низкий"
            grade_color = "#e74c3c"

        structural_percent = structural * 100
        functional_percent = functional * 100
        total_percent = total * 100
        efficiency_percent = efficiency_score * 100 if efficiency_score else 0

        context = {
            'model': model,
            'stats': stats,
            'image': image_base64,
            'structural': structural,
            'functional': functional,
            'total': total,
            'efficiency': efficiency_score,
            'efficiency_percent': efficiency_percent,
            'business_value': business_value,
            'risk_score': risk_score,
            'synergy_factor': synergy_factor,
            'economic_impact': economic_impact,
            'metcalfe_value': metcalfe_value,
            'scale_efficiency': scale_efficiency,
            'interpretations': interpretations,
            'grade': grade,
            'grade_color': grade_color,
            'structural_percent': structural_percent,
            'functional_percent': functional_percent,
            'total_percent': total_percent,
            'potential_growth': (1 - (model.efficiency or 0)) * 100,
        }
        messages.success(request, f'Модель создана! Общий сетевой эффект: {total:.4f}')
        return render(request, 'modeling/result_detail.html', context)

    except Exception as e:
        messages.error(request, f'Ошибка: {str(e)}')
        return redirect('modeling:generate_form')


def view_result(request, pk):
    model = get_object_or_404(GraphModel, pk=pk)
    image_base64 = None
    if model.graph_data:
        try:
            if model.directed == 'directed':
                G = nx.DiGraph()
            else:
                G = nx.Graph()
            G.add_nodes_from(model.graph_data['nodes'])
            for edge in model.graph_data['edges']:
                w = edge.get('weight', 1)
                G.add_edge(edge['source'], edge['target'], weight=w)
            image_base64 = visualize_graph(
                G,
                model.directed == 'directed',
                model.weight_type == 'weighted',
                model.balance_type == 'balanced'
            )
        except Exception as e:
            print(f"Error generating graph: {e}")
            image_base64 = None

    # Дополнительные метрики
    density_pct = model.get_density() * 100
    synergy_factor = model.synergy_factor if model.synergy_factor else model.num_agents
    metcalfe_value = model.num_agents ** 2
    economic_impact = model.economic_impact if model.economic_impact else 0
    potential_growth = (1 - (model.efficiency or 0)) * 100

    structural_percent = (model.structural_effect * 100) if model.structural_effect else 0
    functional_percent = (model.functional_effect * 100) if model.functional_effect else 0
    total_percent = (model.total_effect * 100) if model.total_effect else 0
    efficiency_percent = (model.efficiency * 100) if model.efficiency else 0

    if model.total_effect and model.total_effect >= 60:
        grade = "Высокий"
        grade_color = "#27ae60"
    elif model.total_effect and model.total_effect >= 30:
        grade = "Средний"
        grade_color = "#f39c12"
    else:
        grade = "Низкий"
        grade_color = "#e74c3c"

    interpretations = model.recommendation.split('\n') if model.recommendation else []

    context = {
        'model': model,
        'image': image_base64,
        'stats': {
            'nodes': model.num_agents,
            'edges': model.num_edges,
            'density_percent': round(density_pct, 1),
            'is_directed': model.directed == 'directed',
            'is_weighted': model.weight_type == 'weighted',
            'weight_avg': model.weight_avg if model.weight_avg else 1.0,
        },
        'structural': model.structural_effect or 0,
        'functional': model.functional_effect or 0,
        'total': model.total_effect or 0,
        'efficiency': model.efficiency or 0,
        'business_value': model.business_value or 0,
        'risk_score': model.risk_score or 0,
        'interpretations': interpretations,
        'grade': grade,
        'grade_color': grade_color,
        'synergy_factor': synergy_factor,
        'metcalfe_value': metcalfe_value,
        'scale_efficiency': density_pct,
        'economic_impact': economic_impact,
        'potential_growth': potential_growth,
        'structural_percent': structural_percent,
        'functional_percent': functional_percent,
        'total_percent': total_percent,
        'efficiency_percent': efficiency_percent,
        'scale_recommendation': max(3, int(model.num_agents * 0.2)),
        'risk_recommendation': max(2, int(model.num_agents * 0.1)),
        'monitoring_interval': max(30, int(100 - density_pct)),
    }
    return render(request, 'modeling/result_detail.html', context)


def list_results(request):
    results = GraphModel.objects.all()
    paginator = Paginator(results, 15)
    page = request.GET.get('page', 1)
    page_obj = paginator.get_page(page)
    return render(request, 'modeling/list_results.html', {'page_obj': page_obj})


def delete_result(request, pk):
    model = get_object_or_404(GraphModel, pk=pk)
    model_name = model.name or f"Модель #{model.id}"
    model.delete()
    messages.success(request, f'Модель "{model_name}" удалена')
    return redirect('modeling:list_results')


# ======================== ЭКСПОРТ ОТДЕЛЬНОЙ МОДЕЛИ ========================

# ======================== ЭКСПОРТ ОТДЕЛЬНОЙ МОДЕЛИ (ИСПРАВЛЕННЫЕ ВЕРСИИ) ========================

def export_json(request, pk):
    """Экспорт в JSON с корректными научными формулами"""
    model = get_object_or_404(GraphModel, pk=pk)
    
    # Правильные вычисления
    density = model.get_density()
    max_edges = model.num_agents * (model.num_agents - 1)
    if model.directed == 'undirected':
        max_edges //= 2
    density_percent = density * 100
    
    synergy_factor = model.synergy_factor if model.synergy_factor is not None else model.num_agents
    economic_impact = model.economic_impact if model.economic_impact is not None else ((model.num_agents + 1) ** 2 / (model.num_agents ** 2) - 1) * 100
    
    data = {
        'metadata': {
            'id': model.id,
            'name': model.name,
            'created_at': model.created_at.isoformat(),
            'export_date': datetime.now().isoformat(),
        },
        'parameters': {
            'num_agents': model.num_agents,
            'num_edges': model.num_edges,
            'density': round(density, 4),
            'directed': model.directed,
            'directed_display': model.get_directed_display(),
            'weight_type': model.weight_type,
            'weight_type_display': model.get_weight_type_display(),
            'balance_type': model.balance_type if hasattr(model, 'balance_type') else 'balanced',
            'balance_display': 'Сбалансированный' if (hasattr(model, 'balance_type') and model.balance_type == 'balanced') else 'Несбалансированный',
            'weight_min': model.weight_min,
            'weight_max': model.weight_max,
            'weight_avg': model.weight_avg if model.weight_avg is not None else 1.0,
        },
        'results': {
            'structural_effect': model.structural_effect,
            'functional_effect': model.functional_effect,
            'total_effect': model.total_effect,
            'efficiency': model.efficiency,
            'business_value': model.business_value,
            'risk_score': model.risk_score,
        },
        'synergy_metrics': {
            'synergy_factor': synergy_factor,
            'metcalfe_value': model.num_agents ** 2,
            'scale_efficiency': density_percent,
            'economic_impact': economic_impact,
            'potential_growth': (1 - (model.efficiency or 0)) * 100,
        },
        'statistics': {
            'max_possible_edges': max_edges,
            'density_percent': round(density_percent, 1),
        },
        'recommendations': {
            'scale_recommendation': max(3, int(model.num_agents * 0.2)),
            'risk_recommendation': max(2, int(model.num_agents * 0.1)),
            'monitoring_interval': max(30, int(100 - density_percent)),
        },
        'graph_data': model.graph_data,
        'recommendation_text': model.recommendation,
    }
    
    response = HttpResponse(json.dumps(data, ensure_ascii=False, indent=2), content_type='application/json')
    response['Content-Disposition'] = f'attachment; filename="model_{model.id}_report.json"'
    return response


def export_csv(request, pk):
    """Экспорт в CSV с корректными научными формулами"""
    model = get_object_or_404(GraphModel, pk=pk)
    
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="model_{model.id}_report.csv"'
    writer = csv.writer(response)
    
    # Заголовок
    writer.writerow(['=' * 80])
    writer.writerow(['ОТЧЕТ ПО МОДЕЛИРОВАНИЮ СЕТЕВОГО ЭФФЕКТА'])
    writer.writerow(['=' * 80])
    writer.writerow([])
    
    # Метаданные
    writer.writerow(['МЕТАДАННЫЕ'])
    writer.writerow(['Параметр', 'Значение'])
    writer.writerow(['ID модели', model.id])
    writer.writerow(['Название', model.name or '-'])
    writer.writerow(['Дата создания', model.created_at.strftime('%d.%m.%Y %H:%M:%S')])
    writer.writerow([])
    
    # Параметры модели
    density = model.get_density()
    max_edges = model.num_agents * (model.num_agents - 1)
    if model.directed == 'undirected':
        max_edges //= 2
    density_pct = density * 100
    
    writer.writerow(['ПАРАМЕТРЫ МОДЕЛИ'])
    writer.writerow(['Параметр', 'Значение'])
    writer.writerow(['Количество агентов', model.num_agents])
    writer.writerow(['Количество связей', model.num_edges])
    writer.writerow(['Плотность сети', f"{density:.4f} ({density_pct:.1f}%)"])
    writer.writerow(['Направленность', model.get_directed_display()])
    writer.writerow(['Тип весов', model.get_weight_type_display()])
    writer.writerow(['Сбалансированность', 'Сбалансированный' if (hasattr(model, 'balance_type') and model.balance_type == 'balanced') else 'Несбалансированный'])
    if model.weight_min:
        writer.writerow(['Мин. вес', model.weight_min])
        writer.writerow(['Макс. вес', model.weight_max])
        writer.writerow(['Средний вес', model.weight_avg if model.weight_avg is not None else 1.0])
    writer.writerow([])
    
    # Результаты расчета
    writer.writerow(['РЕЗУЛЬТАТЫ РАСЧЕТА'])
    writer.writerow(['Показатель', 'Значение', 'Процент от максимума'])
    writer.writerow(['Структурный эффект', model.structural_effect or 0, f'{(model.structural_effect or 0) * 100:.1f}%'])
    writer.writerow(['Функциональный эффект', model.functional_effect or 0, f'{(model.functional_effect or 0) * 100:.1f}%'])
    writer.writerow(['Общий сетевой эффект', model.total_effect or 0, f'{(model.total_effect or 0) * 100:.1f}%'])
    writer.writerow(['Эффективность цепочки поставок', model.efficiency or 0, f'{(model.efficiency or 0) * 100:.1f}%'])
    writer.writerow(['Бизнес-ценность', model.business_value or 0, '-'])
    writer.writerow(['Уровень риска', f"{model.risk_score or 0}%", '-'])
    writer.writerow([])
    
    # Бизнес-метрики
    synergy_factor = model.synergy_factor if model.synergy_factor is not None else model.num_agents
    economic_impact = model.economic_impact if model.economic_impact is not None else ((model.num_agents + 1) ** 2 / (model.num_agents ** 2) - 1) * 100
    
    writer.writerow(['БИЗНЕС-МЕТРИКИ'])
    writer.writerow(['Показатель', 'Значение'])
    writer.writerow(['Плотность сети', f'{density_pct:.1f}%'])
    writer.writerow(['Максимально возможное количество связей', max_edges])
    writer.writerow(['Мультипликатор синергии', f'{synergy_factor:.1f}x'])
    writer.writerow(['Ценность сети (n²)', model.num_agents ** 2])
    writer.writerow(['Экономический эффект', f'{economic_impact:.0f}%'])
    writer.writerow(['Потенциал роста', f'{(100 - (model.total_effect or 0)) * 100:.1f}%'])
    writer.writerow([])
    
    # Рекомендации
    writer.writerow(['РЕКОМЕНДАЦИИ'])
    writer.writerow(['Рекомендация'])
    if model.recommendation:
        for line in model.recommendation.split('\n'):
            if line.strip():
                writer.writerow([line])
    else:
        writer.writerow(['Нет доступных рекомендаций'])
    writer.writerow([])
    
    # Структура связей
    if model.graph_data and model.graph_data.get('edges'):
        writer.writerow(['СТРУКТУРА СВЯЗЕЙ'])
        writer.writerow(['Источник', 'Цель', 'Вес'])
        for edge in model.graph_data['edges']:
            writer.writerow([edge['source'], edge['target'], edge.get('weight', 1)])
    
    return response


def export_excel(request, pk):
    """Экспорт в Excel с корректными научными формулами"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    
    model = get_object_or_404(GraphModel, pk=pk)
    wb = openpyxl.Workbook()
    
    # Стили
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Вычисления
    density = model.get_density()
    max_edges = model.num_agents * (model.num_agents - 1)
    if model.directed == 'undirected':
        max_edges //= 2
    density_pct = density * 100
    synergy_factor = model.synergy_factor if model.synergy_factor is not None else model.num_agents
    economic_impact = model.economic_impact if model.economic_impact is not None else ((model.num_agents + 1) ** 2 / (model.num_agents ** 2) - 1) * 100
    potential_growth = (1 - (model.efficiency or 0)) * 100
    
    # Лист 1: Основная информация
    ws1 = wb.active
    ws1.title = "Основная информация"
    ws1.merge_cells('A1:B1')
    ws1['A1'].value = "ОТЧЕТ ПО МОДЕЛИРОВАНИЮ СЕТЕВОГО ЭФФЕКТА"
    ws1['A1'].font = Font(bold=True, size=16)
    ws1['A1'].alignment = Alignment(horizontal='center')
    ws1['A3'] = f"Модель: {model.name or 'Без названия'}"
    ws1['A3'].font = Font(bold=True, size=12)
    ws1['A4'] = f"ID: {model.id}"
    ws1['A5'] = f"Дата создания: {model.created_at.strftime('%d.%m.%Y %H:%M:%S')}"
    
    # Параметры модели
    row = 7
    ws1[f'A{row}'] = "ПАРАМЕТРЫ МОДЕЛИ"
    ws1[f'A{row}'].font = Font(bold=True, size=12, color="3498db")
    row += 1
    
    params = [
        ('Количество агентов', model.num_agents),
        ('Количество связей', f"{model.num_edges} (максимум: {max_edges})"),
        ('Плотность сети', f"{density:.4f} ({density_pct:.1f}%)"),
        ('Направленность', model.get_directed_display()),
        ('Тип весов', model.get_weight_type_display()),
        ('Сбалансированность', 'Сбалансированный' if (hasattr(model, 'balance_type') and model.balance_type == 'balanced') else 'Несбалансированный'),
    ]
    if model.weight_min:
        params.append(('Мин. вес', model.weight_min))
        params.append(('Макс. вес', model.weight_max))
        params.append(('Средний вес', model.weight_avg if model.weight_avg is not None else 1.0))
    
    for param, value in params:
        ws1[f'A{row}'] = param
        ws1[f'A{row}'].font = Font(bold=True)
        ws1[f'B{row}'] = value
        row += 1
    
    # Результаты
    row += 1
    ws1[f'A{row}'] = "РЕЗУЛЬТАТЫ РАСЧЕТА"
    ws1[f'A{row}'].font = Font(bold=True, size=12, color="3498db")
    row += 1
    
    results = [
        ('Структурный эффект', model.structural_effect or 0, f'{(model.structural_effect or 0) * 100:.1f}%'),
        ('Функциональный эффект', model.functional_effect or 0, f'{(model.functional_effect or 0) * 100:.1f}%'),
        ('Общий сетевой эффект', model.total_effect or 0, f'{(model.total_effect or 0) * 100:.1f}%'),
        ('Эффективность цепочки поставок', model.efficiency or 0, f'{(model.efficiency or 0) * 100:.1f}%'),
        ('Бизнес-ценность', model.business_value or 0, '-'),
        ('Уровень риска', f"{model.risk_score or 0}%", '-'),
    ]
    
    for result in results:
        ws1[f'A{row}'] = result[0]
        ws1[f'A{row}'].font = Font(bold=True)
        ws1[f'B{row}'] = result[1]
        ws1[f'C{row}'] = result[2]
        row += 1
    
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 25
    ws1.column_dimensions['C'].width = 20
    
    # Лист 2: Бизнес-аналитика
    ws2 = wb.create_sheet("Бизнес-аналитика")
    business_data = [
        ('Плотность сети', f'{density_pct:.1f}%', 'Чем выше плотность, тем эффективнее сеть'),
        ('Максимально возможное количество связей', max_edges, 'При текущем количестве агентов'),
        ('Мультипликатор синергии', f'{synergy_factor:.1f}x', 'Эффект масштабирования (закон Меткалфа)'),
        ('Ценность сети (n²)', model.num_agents ** 2, 'По закону Меткалфа'),
        ('Экономический эффект', f'{economic_impact:.0f}%', 'Прирост при добавлении одного агента'),
        ('Потенциал роста', f'{potential_growth:.1f}%', 'До достижения максимума'),
    ]
    for i, (label, val, desc) in enumerate(business_data, 1):
        ws2[f'A{i}'] = label
        ws2[f'A{i}'].font = Font(bold=True)
        ws2[f'B{i}'] = val
        ws2[f'C{i}'] = desc
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 25
    ws2.column_dimensions['C'].width = 40
    
    # Лист 3: Рекомендации
    ws3 = wb.create_sheet("Рекомендации")
    ws3['A1'] = "РЕКОМЕНДАЦИИ ДЛЯ БИЗНЕСА"
    ws3['A1'].font = Font(bold=True, size=14, color="f39c12")
    row = 3
    if model.recommendation:
        for line in model.recommendation.split('\n'):
            if line.strip():
                ws3[f'A{row}'] = line
                row += 1
    else:
        ws3['A3'] = "Нет доступных рекомендаций"
    ws3.column_dimensions['A'].width = 80
    
    # Лист 4: Структура связей
    if model.graph_data and model.graph_data.get('edges'):
        ws4 = wb.create_sheet("Структура связей")
        headers = ['Источник', 'Цель', 'Вес']
        for col, header in enumerate(headers, 1):
            cell = ws4.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
        for r, edge in enumerate(model.graph_data['edges'], 2):
            ws4.cell(row=r, column=1, value=edge['source'])
            ws4.cell(row=r, column=2, value=edge['target'])
            ws4.cell(row=r, column=3, value=edge.get('weight', 1))
        ws4.column_dimensions['A'].width = 15
        ws4.column_dimensions['B'].width = 15
        ws4.column_dimensions['C'].width = 15
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="model_{model.id}_report.xlsx"'
    wb.save(response)
    return response


def export_pdf(request, pk):
    """Экспорт результата модели в HTML отчет с корректными формулами и визуализацией"""
    from datetime import datetime
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import base64
    from io import BytesIO
    import networkx as nx
    from .views import visualize_graph  # импортируем улучшенную функцию из views
    
    model = get_object_or_404(GraphModel, pk=pk)
    
    # Генерируем изображение графа с учётом сбалансированности
    graph_image = None
    if model.graph_data:
        try:
            if model.directed == 'directed':
                G = nx.DiGraph()
            else:
                G = nx.Graph()
            G.add_nodes_from(model.graph_data['nodes'])
            for edge in model.graph_data['edges']:
                w = edge.get('weight', 1)
                G.add_edge(edge['source'], edge['target'], weight=w)
            graph_image = visualize_graph(
                G,
                model.directed == 'directed',
                model.weight_type == 'weighted',
                model.balance_type == 'balanced'
            )
        except Exception as e:
            print(f"Error generating graph: {e}")
            graph_image = None
    
    # Правильные расчёты
    density = model.get_density()
    density_pct = density * 100
    max_edges = model.num_agents * (model.num_agents - 1)
    if model.directed == 'undirected':
        max_edges //= 2
    
    synergy_factor = model.synergy_factor if model.synergy_factor is not None else model.num_agents
    metcalfe_value = model.num_agents ** 2
    economic_impact = model.economic_impact if model.economic_impact is not None else ((model.num_agents + 1) ** 2 / (model.num_agents ** 2) - 1) * 100
    potential_growth = (1 - (model.efficiency or 0)) * 100
    
    structural_percent = (model.structural_effect * 100) if model.structural_effect else 0
    functional_percent = (model.functional_effect * 100) if model.functional_effect else 0
    total_percent = (model.total_effect * 100) if model.total_effect else 0
    efficiency_percent = (model.efficiency * 100) if model.efficiency else 0
    
    scale_recommendation = max(3, int(model.num_agents * 0.2))
    risk_recommendation = max(2, int(model.num_agents * 0.1))
    monitoring_interval = max(30, int(100 - density_pct))
    
    structural_value = f"{model.structural_effect:.4f}" if model.structural_effect else "0"
    functional_value = f"{model.functional_effect:.4f}" if model.functional_effect else "0"
    total_value = f"{model.total_effect:.4f}" if model.total_effect else "0"
    efficiency_value = f"{model.efficiency:.4f}" if model.efficiency else "0"
    risk_value = f"{model.risk_score:.1f}" if model.risk_score else "0"
    
    if model.total_effect and model.total_effect >= 60:
        grade = "ВЫСОКИЙ"
        grade_color = "#27ae60"
    elif model.total_effect and model.total_effect >= 30:
        grade = "СРЕДНИЙ"
        grade_color = "#f39c12"
    else:
        grade = "НИЗКИЙ"
        grade_color = "#e74c3c"
    
    interpretations_html = ""
    if model.recommendation:
        for line in model.recommendation.split('\n'):
            if line.strip():
                interpretations_html += f"<li>{line}</li>"
    
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>Отчет по модели #{model.id}</title>
        <style>
            @page {{ size: A4; margin: 1.5cm; }}
            body {{ font-family: 'Segoe UI', Arial, sans-serif; margin: 0; padding: 20px; color: #2c3e50; line-height: 1.5; }}
            .header-card {{ background: linear-gradient(135deg, #ffffff 0%, #f8f9fa 100%); border-radius: 28px; padding: 24px; margin-bottom: 24px; border: 1px solid rgba(52,152,219,0.15); }}
            .header-title {{ font-size: 28px; color: #2c3e50; margin-bottom: 8px; }}
            .date {{ text-align: right; color: #7f8c8d; font-size: 11px; margin-bottom: 20px; }}
            .grade-badge {{ background: {grade_color}15; padding: 8px 20px; border-radius: 40px; border: 2px solid {grade_color}; display: inline-block; }}
            .grade-text {{ color: {grade_color}; font-weight: bold; }}
            .metrics-grid {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: 20px; margin-bottom: 24px; }}
            .metric-card {{ border-radius: 24px; padding: 24px; text-align: center; color: white; }}
            .metric-structural {{ background: linear-gradient(135deg, #2c3e50 0%, #3498db 100%); }}
            .metric-functional {{ background: linear-gradient(135deg, #27ae60 0%, #2ecc71 100%); }}
            .metric-total {{ background: linear-gradient(135deg, #8e44ad 0%, #9b59b6 100%); }}
            .metric-value {{ font-size: 38px; font-weight: bold; margin: 12px 0; }}
            .progress-bar {{ height: 6px; background: rgba(255,255,255,0.25); border-radius: 4px; overflow: hidden; margin-top: 12px; }}
            .progress-fill {{ height: 100%; background: white; border-radius: 4px; }}
            .risk-efficiency-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 20px; margin-bottom: 24px; }}
            .info-card {{ border-radius: 24px; padding: 24px; color: white; }}
            .risk-card {{ background: linear-gradient(135deg, #c0392b 0%, #e74c3c 100%); }}
            .efficiency-card {{ background: linear-gradient(135deg, #2471a3 0%, #3498db 100%); }}
            .info-value {{ font-size: 48px; font-weight: bold; margin: 12px 0; }}
            .synergy-card {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); border-radius: 24px; padding: 24px; margin-bottom: 24px; color: white; }}
            .synergy-title {{ text-align: center; font-size: 20px; font-weight: bold; margin-bottom: 20px; }}
            .synergy-grid {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 15px; text-align: center; }}
            .synergy-item {{ background: rgba(255,255,255,0.15); border-radius: 16px; padding: 15px; }}
            .synergy-value {{ font-size: 28px; font-weight: bold; }}
            .params-card {{ background: white; border-radius: 24px; padding: 24px; margin-bottom: 24px; box-shadow: 0 4px 15px rgba(0,0,0,0.06); }}
            .params-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 15px; }}
            .param-box {{ background: #f8f9fa; border-radius: 16px; padding: 15px; border-left: 4px solid #3498db; }}
            .visualization-card {{ background: white; border-radius: 24px; padding: 24px; margin-bottom: 24px; text-align: center; }}
            .graph-image {{ max-width: 100%; border-radius: 16px; }}
            .interpretation-card {{ background: #f0f7ff; border-radius: 24px; padding: 24px; margin-bottom: 24px; border-left: 5px solid #3498db; }}
            .interpretation-list {{ list-style: none; padding: 0; }}
            .interpretation-list li {{ margin-bottom: 10px; padding-left: 24px; position: relative; }}
            .interpretation-list li::before {{ content: "▸"; position: absolute; left: 0; color: #3498db; }}
            .recommendations-card {{ background: #fffbf0; border-radius: 24px; padding: 24px; margin-bottom: 24px; border-left: 5px solid #f39c12; }}
            .footer {{ text-align: center; margin-top: 30px; padding-top: 15px; border-top: 1px solid #e0e0e0; font-size: 10px; color: #7f8c8d; }}
        </style>
    </head>
    <body>
        <div class="date">Дата отчета: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}</div>
        <div class="header-card">
            <div style="display: flex; justify-content: space-between; align-items: center;">
                <div><div class="header-title">📊 {model.name or 'Модель'}</div><div>ID: {model.id} | Создано: {model.created_at.strftime('%d.%m.%Y %H:%M:%S')}</div></div>
                <div class="grade-badge"><span class="grade-text">Уровень: {grade}</span></div>
            </div>
        </div>
        <div class="metrics-grid">
            <div class="metric-card metric-structural"><div class="metric-value">{structural_value}</div><div>Структурный эффект</div><div class="progress-bar"><div class="progress-fill" style="width:{structural_percent:.0f}%"></div></div></div>
            <div class="metric-card metric-functional"><div class="metric-value">{functional_value}</div><div>Функциональный эффект</div><div class="progress-bar"><div class="progress-fill" style="width:{functional_percent:.0f}%"></div></div></div>
            <div class="metric-card metric-total"><div class="metric-value">{total_value}</div><div>Общий сетевой эффект</div><div class="progress-bar"><div class="progress-fill" style="width:{total_percent:.0f}%"></div></div></div>
        </div>
        <div class="risk-efficiency-grid">
            <div class="info-card risk-card"><h3>⚠️ Оценка рисков</h3><div class="info-value">{risk_value}%</div><div>Уровень риска</div><div class="progress-bar"><div class="progress-fill" style="width:{risk_value}%"></div></div><div style="margin-top:16px;">{'🔴 Высокий риск' if model.risk_score and model.risk_score > 30 else '🟢 Риск в пределах'}</div></div>
            <div class="info-card efficiency-card"><h3>📦 Эффективность цепочки поставок</h3><div class="info-value">{efficiency_value}</div><div>Общая эффективность</div><div class="progress-bar"><div class="progress-fill" style="width:{efficiency_percent:.0f}%"></div></div><div style="margin-top:16px;">📊 Плотность: {density_pct:.1f}%</div></div>
        </div>
        <div class="synergy-card">
            <div class="synergy-title">🚀 Синергия эффекта от масштаба</div>
            <div class="synergy-grid">
                <div class="synergy-item"><div class="synergy-value">{synergy_factor:.2f}x</div><div>Мультипликатор</div></div>
                <div class="synergy-item"><div class="synergy-value">{metcalfe_value}</div><div>Ценность сети (n²)</div></div>
                <div class="synergy-item"><div class="synergy-value">{density_pct:.1f}%</div><div>Плотность</div></div>
                <div class="synergy-item"><div class="synergy-value">{economic_impact:.0f}%</div><div>Экономический эффект</div></div>
            </div>
            <div class="synergy-footer">📈 При масштабировании на 15% прирост до {economic_impact:.0f}%</div>
        </div>
        <div class="params-card">
            <div class="params-title">⚙️ Параметры модели</div>
            <div class="params-grid">
                <div class="param-box"><strong>Агентов</strong><span>{model.num_agents}</span></div>
                <div class="param-box"><strong>Связей</strong><span>{model.num_edges}</span></div>
                <div class="param-box"><strong>Плотность</strong><span>{density_pct:.1f}%</span></div>
                <div class="param-box"><strong>Направленность</strong><span>{model.get_directed_display()}</span></div>
                <div class="param-box"><strong>Тип весов</strong><span>{model.get_weight_type_display()}</span></div>
                <div class="param-box"><strong>Сбалансированность</strong><span>{'Сбалансированный' if model.balance_type == 'balanced' else 'Несбалансированный'}</span></div>
            </div>
        </div>
        {f'<div class="visualization-card"><h3>Визуализация сети</h3><img src="data:image/png;base64,{graph_image}" class="graph-image"></div>' if graph_image else ''}
        <div class="interpretation-card"><h3>Бизнес-интерпретация</h3><ul class="interpretation-list">{interpretations_html}</ul></div>
        <div class="recommendations-card"><h3>Рекомендации</h3><ul><li>📈 Увеличьте плотность связей — текущий уровень {density_pct:.0f}%</li><li>🛡️ Внедрите {risk_recommendation} резервных каналов</li><li>🚀 Добавьте {scale_recommendation} новых агентов</li><li>📊 Мониторинг каждые {monitoring_interval} дней</li></ul></div>
        <div class="footer"><p>© 2026 NetEffect Modeler. Все права защищены.</p></div>
    </body>
    </html>
    """
    response = HttpResponse(html_content, content_type='text/html')
    response['Content-Disposition'] = f'attachment; filename="model_{model.id}_report.html"'
    return response


# ======================== ЭКСПОРТ ДАШБОРДА (ИСПРАВЛЕННЫЕ ВЕРСИИ) ========================

def export_dashboard_json(request):
    """Экспорт дашборда в JSON с корректными научными полями"""
    models = GraphModel.objects.all()
    data = []
    for model in models:
        data.append({
            'id': model.id,
            'name': model.name,
            'created_at': model.created_at.isoformat(),
            'num_agents': model.num_agents,
            'num_edges': model.num_edges,
            'density': model.get_density(),
            'directed': model.directed,
            'weight_type': model.weight_type,
            'balance_type': model.balance_type,
            'structural_effect': model.structural_effect,
            'functional_effect': model.functional_effect,
            'total_effect': model.total_effect,
            'efficiency': model.efficiency,
            'business_value': model.business_value,
            'risk_score': model.risk_score,
        })
    response = HttpResponse(json.dumps(data, ensure_ascii=False, indent=2), content_type='application/json')
    response['Content-Disposition'] = 'attachment; filename="dashboard_export.json"'
    return response


def export_dashboard_csv(request):
    """Экспорт дашборда в CSV с корректными научными полями"""
    models = GraphModel.objects.all()
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="dashboard_export.csv"'
    writer = csv.writer(response)
    writer.writerow([
        'ID', 'Название', 'Дата', 'Агентов', 'Связей', 'Плотность',
        'Направленность', 'Тип весов', 'Сбалансированность',
        'Структурный эффект', 'Функциональный эффект', 'Общий эффект',
        'Эффективность', 'Бизнес-ценность', 'Уровень риска'
    ])
    for model in models:
        writer.writerow([
            model.id,
            model.name or '-',
            model.created_at.strftime('%d.%m.%Y %H:%M:%S'),
            model.num_agents,
            model.num_edges,
            f"{model.get_density():.4f}",
            model.get_directed_display(),
            model.get_weight_type_display(),
            model.balance_type if hasattr(model, 'balance_type') else 'balanced',
            model.structural_effect or 0,
            model.functional_effect or 0,
            model.total_effect or 0,
            model.efficiency or 0,
            model.business_value or 0,
            model.risk_score or 0,
        ])
    return response


def export_dashboard_excel(request):
    """Экспорт дашборда в Excel с корректными научными полями"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime

    models = GraphModel.objects.all().order_by('-created_at')
    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Лист 1: Основная информация
    ws1 = wb.active
    ws1.title = "Основная информация"
    ws1.merge_cells('A1:D1')
    cell = ws1['A1']
    cell.value = "ОТЧЕТ ПО ДАШБОРДУ МОДЕЛИРОВАНИЯ СЕТЕВЫХ ЭФФЕКТОВ"
    cell.font = Font(bold=True, size=16)
    cell.alignment = Alignment(horizontal='center')
    ws1['A3'] = f"Дата формирования: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}"
    ws1['A3'].font = Font(italic=True)
    ws1['A5'] = "СТАТИСТИКА"
    ws1['A5'].font = Font(bold=True, size=12)
    ws1['A5'].fill = PatternFill(start_color="e8f4f8", end_color="e8f4f8", fill_type="solid")

    stats_data = [
        ("Всего моделей", models.count()),
        ("Направленных графов", models.filter(directed='directed').count()),
        ("Ненаправленных графов", models.filter(directed='undirected').count()),
        ("Взвешенных графов", models.filter(weight_type='weighted').count()),
        ("Без весов", models.filter(weight_type='unweighted').count()),
    ]
    row = 6
    for stat in stats_data:
        ws1[f'A{row}'] = stat[0]
        ws1[f'A{row}'].font = Font(bold=True)
        ws1[f'B{row}'] = stat[1]
        row += 1

    # Лист 2: Все модели
    ws2 = wb.create_sheet("Все модели")
    headers = [
        'ID', 'Название', 'Дата', 'Агентов', 'Связей', 'Плотность',
        'Направленность', 'Тип весов', 'Сбалансированность',
        'Структурный эффект', 'Функциональный эффект', 'Общий эффект',
        'Эффективность', 'Бизнес-ценность', 'Уровень риска'
    ]
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    for row, model in enumerate(models, 2):
        ws2.cell(row=row, column=1, value=model.id).border = thin_border
        ws2.cell(row=row, column=2, value=model.name or '-').border = thin_border
        ws2.cell(row=row, column=3, value=model.created_at.strftime('%d.%m.%Y %H:%M')).border = thin_border
        ws2.cell(row=row, column=4, value=model.num_agents).border = thin_border
        ws2.cell(row=row, column=5, value=model.num_edges).border = thin_border
        ws2.cell(row=row, column=6, value=round(model.get_density(), 4)).border = thin_border
        ws2.cell(row=row, column=7, value=model.get_directed_display()).border = thin_border
        ws2.cell(row=row, column=8, value=model.get_weight_type_display()).border = thin_border
        ws2.cell(row=row, column=9, value=model.balance_type if hasattr(model, 'balance_type') else 'balanced').border = thin_border
        ws2.cell(row=row, column=10, value=model.structural_effect or 0).border = thin_border
        ws2.cell(row=row, column=11, value=model.functional_effect or 0).border = thin_border
        ws2.cell(row=row, column=12, value=model.total_effect or 0).border = thin_border
        ws2.cell(row=row, column=13, value=model.efficiency or 0).border = thin_border
        ws2.cell(row=row, column=14, value=model.business_value or 0).border = thin_border
        ws2.cell(row=row, column=15, value=model.risk_score or 0).border = thin_border

    for col in range(1, len(headers) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 15
    ws2.column_dimensions['B'].width = 25
    ws2.column_dimensions['C'].width = 18

    # Лист 3: Топ моделей
    ws3 = wb.create_sheet("Топ моделей")
    top_models = models.order_by('-total_effect')[:10]
    ws3['A1'] = "ТОП-10 МОДЕЛЕЙ ПО ОБЩЕМУ СЕТЕВОМУ ЭФФЕКТУ"
    ws3['A1'].font = Font(bold=True, size=14)
    headers3 = ['Место', 'ID', 'Название', 'Общий эффект', 'Структурный', 'Функциональный', 'Агентов']
    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    for idx, model in enumerate(top_models, 1):
        row = 3 + idx
        ws3.cell(row=row, column=1, value=idx).border = thin_border
        ws3.cell(row=row, column=2, value=model.id).border = thin_border
        ws3.cell(row=row, column=3, value=model.name or '-').border = thin_border
        ws3.cell(row=row, column=4, value=model.total_effect or 0).border = thin_border
        ws3.cell(row=row, column=5, value=model.structural_effect or 0).border = thin_border
        ws3.cell(row=row, column=6, value=model.functional_effect or 0).border = thin_border
        ws3.cell(row=row, column=7, value=model.num_agents).border = thin_border
    for col in range(1, 8):
        ws3.column_dimensions[get_column_letter(col)].width = 15

    # Лист 4: Статистическая сводка
    ws4 = wb.create_sheet("Статистическая сводка")
    effects = [m.total_effect for m in models if m.total_effect is not None]
    if effects:
        import numpy as np
        stats_list = [
            ("Средний сетевой эффект", np.mean(effects)),
            ("Медианный эффект", np.median(effects)),
            ("Минимальный эффект", np.min(effects)),
            ("Максимальный эффект", np.max(effects)),
            ("Стандартное отклонение", np.std(effects)),
            ("25-й перцентиль", np.percentile(effects, 25)),
            ("75-й перцентиль", np.percentile(effects, 75)),
        ]
    else:
        stats_list = []
    ws4['A1'] = "СТАТИСТИЧЕСКАЯ СВОДКА"
    ws4['A1'].font = Font(bold=True, size=14)
    for idx, (name, value) in enumerate(stats_list, 3):
        ws4[f'A{idx}'] = name
        ws4[f'A{idx}'].font = Font(bold=True)
        ws4[f'B{idx}'] = round(value, 4) if isinstance(value, float) else value
    ws4.column_dimensions['A'].width = 30
    ws4.column_dimensions['B'].width = 20

    # Лист 5: Распределение
    ws5 = wb.create_sheet("Распределение")
    high = models.filter(total_effect__gte=60).count()
    medium = models.filter(total_effect__gte=30, total_effect__lt=60).count()
    low = models.filter(total_effect__lt=30).count()
    ws5['A1'] = "РАСПРЕДЕЛЕНИЕ МОДЕЛЕЙ ПО УРОВНЮ ЭФФЕКТИВНОСТИ"
    ws5['A1'].font = Font(bold=True, size=14)
    distribution = [
        ("Высокий (≥60)", high, f"{high/models.count()*100:.1f}%" if models.count() > 0 else "0%"),
        ("Средний (30-60)", medium, f"{medium/models.count()*100:.1f}%" if models.count() > 0 else "0%"),
        ("Низкий (<30)", low, f"{low/models.count()*100:.1f}%" if models.count() > 0 else "0%"),
    ]
    for idx, (level, count, percent) in enumerate(distribution, 3):
        ws5[f'A{idx}'] = level
        ws5[f'B{idx}'] = count
        ws5[f'C{idx}'] = percent
    ws5.column_dimensions['A'].width = 25
    ws5.column_dimensions['B'].width = 15
    ws5.column_dimensions['C'].width = 15

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="dashboard_export.xlsx"'
    wb.save(response)
    return response


def export_dashboard_pdf(request):
    """Экспорт дашборда в PDF - HTML отчет (как в модели)"""
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import base64
    from io import BytesIO
    import numpy as np
    from scipy import stats
    from django.db.models import Avg
    from datetime import datetime
    
    models = GraphModel.objects.all()
    total_models = models.count()
    avg_effect = models.filter(total_effect__isnull=False).aggregate(Avg('total_effect'))['total_effect__avg'] or 0
    directed_count = models.filter(directed='directed').count()
    weighted_count = models.filter(weight_type='weighted').count()
    
    # Дополнительная статистика
    avg_structural = models.filter(structural_effect__isnull=False).aggregate(Avg('structural_effect'))['structural_effect__avg'] or 0
    avg_functional = models.filter(functional_effect__isnull=False).aggregate(Avg('functional_effect'))['functional_effect__avg'] or 0
    
    # Сбор данных для графиков
    density_list = []
    effect_list = []
    agents_list = []
    weight_list = []
    weight_effect_list = []
    
    for m in models:
        if m.total_effect is not None:
            # ИСПРАВЛЕНО: m.density -> m.get_density()
            density_list.append(m.get_density())
            effect_list.append(m.total_effect)
            agents_list.append(m.num_agents)
            if m.weight_type == 'weighted':
                if m.weight_min is not None and m.weight_max is not None:
                    weight_list.append((m.weight_min + m.weight_max) / 2)
                    weight_effect_list.append(m.total_effect)
    
    # Корреляции
    density_corr = np.corrcoef(density_list, effect_list)[0, 1] if len(density_list) > 1 else 0
    agents_corr = np.corrcoef(agents_list, effect_list)[0, 1] if len(agents_list) > 1 else 0
    weight_corr = np.corrcoef(weight_list, weight_effect_list)[0, 1] if len(weight_list) > 1 else 0
    
    # Регрессии
    if len(density_list) > 1:
        slope, intercept, r, p, se = stats.linregress(density_list, effect_list)
        density_slope, density_intercept = slope, intercept
    else:
        density_slope, density_intercept = 0, 0
    
    if len(agents_list) > 1:
        slope, intercept, r, p, se = stats.linregress(agents_list, effect_list)
        agents_slope, agents_intercept = slope, intercept
    else:
        agents_slope, agents_intercept = 0, 0
    
    if len(weight_list) > 1:
        slope, intercept, r, p, se = stats.linregress(weight_list, weight_effect_list)
        weight_slope, weight_intercept = slope, intercept
    else:
        weight_slope, weight_intercept = 0, 0
    
    # Статистика для сводки
    effect_array = np.array(effect_list) if effect_list else [0]
    median_effect = np.median(effect_array)
    min_effect = np.min(effect_array) if len(effect_array) > 0 else 0
    max_effect = np.max(effect_array) if len(effect_array) > 0 else 0
    std_effect = np.std(effect_array)
    percentile_25 = np.percentile(effect_array, 25) if len(effect_array) > 0 else 0
    percentile_75 = np.percentile(effect_array, 75) if len(effect_array) > 0 else 0
    growth_potential = max_effect - min_effect
    growth_percent = (growth_potential / max_effect * 100) if max_effect > 0 else 0
    
    # Функция создания графиков
    def create_chart(data, title, xlabel, ylabel, slope, intercept, correlation):
        fig, ax = plt.subplots(figsize=(8, 5))
        if data:
            x_vals = [d[0] for d in data]
            y_vals = [d[1] for d in data]
            colors = plt.cm.Set3(np.linspace(0, 1, len(x_vals)))
            for i, (x, y) in enumerate(zip(x_vals, y_vals)):
                ax.scatter(x, y, color=colors[i], s=80, alpha=0.7, edgecolors='white', linewidth=1.5)
            if len(x_vals) > 1:
                x_line = np.linspace(0, max(x_vals) * 1.1, 100)
                y_line = slope * x_line + intercept
                ax.plot(x_line, y_line, 'r--', linewidth=2, alpha=0.8, label=f'Линия тренда (r={correlation:.3f})')
        ax.set_xlabel(xlabel, fontsize=12, fontweight='bold')
        ax.set_ylabel(ylabel, fontsize=12, fontweight='bold')
        ax.set_title(title, fontsize=14, fontweight='bold', pad=15)
        ax.grid(True, alpha=0.3)
        if data and len(data) > 1:
            ax.legend(loc='lower right')
        ax.set_facecolor('#f8f9fa')
        fig.patch.set_facecolor('white')
        buffer = BytesIO()
        plt.savefig(buffer, format='png', dpi=120, bbox_inches='tight', facecolor='white')
        buffer.seek(0)
        image_base64 = base64.b64encode(buffer.getvalue()).decode()
        plt.close()
        return image_base64
    
    # Подготовка данных
    density_data = list(zip(density_list, effect_list))
    agents_data = list(zip(agents_list, effect_list))
    weight_data = list(zip(weight_list, weight_effect_list))
    
    # Создание графиков
    density_chart = create_chart(density_data, 'Влияние плотности связей на сетевой эффект', 'Количество связей', 'Сетевой эффект', density_slope, density_intercept, density_corr)
    agents_chart = create_chart(agents_data, 'Влияние числа агентов на сетевой эффект', 'Количество агентов', 'Сетевой эффект', agents_slope, agents_intercept, agents_corr)
    
    # ВАЖНО: создаем график для весов, только если есть данные
    if len(weight_data) > 0:
        weight_chart = create_chart(weight_data, 'Влияние весов связей на сетевой эффект', 'Средний вес связи', 'Сетевой эффект', weight_slope, weight_intercept, weight_corr)
        weight_chart_html = f"""
        <div class="chart-card">
            <div class="chart-title">⚖️ Влияние весов связей на сетевой эффект</div>
            <div class="chart-description">
                <p><strong>🎯 Что показывает этот график:</strong> Вес связи отражает интенсивность взаимодействия между агентами. Чем выше вес, тем сильнее связь.</p>
                <p><strong>📈 Интерпретация точек:</strong> Каждая точка — взвешенная модель. Если точки образуют восходящий тренд, это означает, что укрепление существующих связей эффективно.</p>
            </div>
            <div class="chart-container">
                <img src="data:image/png;base64,{weight_chart}" alt="График весов связей">
            </div>
            <div class="stat-card">
                <p><strong>📊 СТАТИСТИЧЕСКАЯ ИНТЕРПРЕТАЦИЯ:</strong></p>
                <p>• <strong>Коэффициент корреляции Пирсона:</strong> 
                    <span class="correlation-value {'correlation-strong' if weight_corr > 0.7 else 'correlation-moderate' if weight_corr > 0.4 else 'correlation-weak'}">
                        {weight_corr:.3f}
                    </span>
                    {'— <span class="correlation-strong" style="background: none; padding: 0;">сильная положительная связь</span>' if weight_corr > 0.7 else '— <span class="correlation-moderate" style="background: none; padding: 0;">умеренная положительная связь</span>' if weight_corr > 0.4 else '— <span class="correlation-weak" style="background: none; padding: 0;">слабая связь</span>'}
                </p>
                <p>• <strong>Уравнение регрессии:</strong></p>
                <div class="regression-formula">
                    🧮 Сетевой эффект = <span style="color: #e67e22; font-weight: bold;">{weight_slope:.6f}</span> × (Вес связи) + <span style="color: #e67e22; font-weight: bold;">{weight_intercept:.4f}</span>
                </div>
                <p>• <strong>Вывод:</strong> 
                {'<span class="conclusion-strong">📈 Положительная зависимость подтверждает, что увеличение интенсивности взаимодействия повышает общую эффективность. Рекомендуется фокусироваться на ключевых связях.</span>' if weight_corr > 0.4 else '<span class="conclusion-weak">📉 Слабая зависимость указывает, что важнее количество связей.</span>'}
                </p>
            </div>
        </div>
        """
    else:
        weight_chart_html = """
        <div class="chart-card">
            <div class="chart-title">⚖️ Влияние весов связей на сетевой эффект</div>
            <div class="chart-description">
                <p><strong>🎯 Что показывает этот график:</strong> Вес связи отражает интенсивность взаимодействия между агентами. Чем выше вес, тем сильнее связь.</p>
            </div>
            <div style="padding: 60px; text-align: center; color: #7f8c8d; background: #f8f9fa; border-radius: 12px;">
                <p>📊 Нет данных для взвешенных графов</p>
                <p style="font-size: 12px; margin-top: 8px;">Создайте взвешенную модель для отображения графика</p>
            </div>
        </div>
        """
    
    # Подготовка значений
    avg_effect_value = f"{avg_effect:.4f}" if avg_effect else "0"
    avg_structural_value = f"{avg_structural:.4f}" if avg_structural else "0"
    avg_functional_value = f"{avg_functional:.4f}" if avg_functional else "0"
    median_effect_value = f"{median_effect:.4f}" if median_effect else "0"
    min_effect_value = f"{min_effect:.4f}" if min_effect else "0"
    max_effect_value = f"{max_effect:.4f}" if max_effect else "0"
    std_effect_value = f"{std_effect:.4f}" if std_effect else "0"
    
    # Последние модели для таблицы
    recent_models = models.order_by('-created_at')[:10]
    
    # Формирование HTML таблицы последних моделей
    models_table_html = ""
    for model in recent_models:
        if model.total_effect and model.total_effect >= 60:
            effect_class = "effect-high"
        elif model.total_effect and model.total_effect >= 30:
            effect_class = "effect-medium"
        else:
            effect_class = "effect-low"
        
        effect_value = f"{model.total_effect:.4f}" if model.total_effect else "N/A"
        
        # ИСПРАВЛЕНО: model.density -> model.get_density()
        density_val = model.get_density()
        
        models_table_html += f"""
        <tr>
            <td>#{model.id}</td>
            <td>{model.name or '-'}</td>
            <td>{model.num_agents}</td>
            <td>{density_val:.3f}</td>
            <td>{model.get_directed_display()} / {model.get_weight_type_display()}</td>
            <td><span class="{effect_class}">{effect_value}</span></td>
            <td><a href="#" style="color: #3498db;">Подробнее →</a></td>
        </tr>
        """
    
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>Бизнес-дашборд - Полный отчет</title>
        <style>
            @page {{
                size: A4;
                margin: 1.5cm;
            }}
            body {{
                font-family: 'Segoe UI', Arial, sans-serif;
                margin: 0;
                padding: 20px;
                color: #2c3e50;
                line-height: 1.5;
            }}
            
            /* KPI карточки */
            .kpi-grid {{
                display: grid;
                grid-template-columns: repeat(4, 1fr);
                gap: 20px;
                margin-bottom: 30px;
            }}
            .kpi-card {{
                border-radius: 20px;
                padding: 24px;
                text-align: center;
                color: white;
            }}
            .kpi-card:nth-child(1) {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); }}
            .kpi-card:nth-child(2) {{ background: linear-gradient(135deg, #11998e 0%, #38ef7d 100%); }}
            .kpi-card:nth-child(3) {{ background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%); }}
            .kpi-card:nth-child(4) {{ background: linear-gradient(135deg, #4facfe 0%, #00f2fe 100%); }}
            .kpi-value {{ font-size: 36px; font-weight: bold; }}
            .kpi-label {{ font-size: 14px; font-weight: 600; letter-spacing: 1px; text-transform: uppercase; margin-top: 8px; }}
            .kpi-trend {{ font-size: 12px; margin-top: 8px; padding: 4px 8px; border-radius: 20px; display: inline-block; background: rgba(255,255,255,0.2); }}
            
            /* Карточки графиков */
            .chart-card {{
                background: white;
                border-radius: 20px;
                padding: 24px;
                margin-bottom: 30px;
                box-shadow: 0 4px 20px rgba(0,0,0,0.08);
            }}
            .chart-title {{
                font-size: 20px;
                font-weight: bold;
                color: #2c3e50;
                margin-bottom: 20px;
                padding-bottom: 12px;
                border-bottom: 3px solid #3498db;
                display: inline-block;
            }}
            .chart-description {{
                background: linear-gradient(135deg, #f0f7ff 0%, #e8f4f8 100%);
                border-radius: 12px;
                padding: 16px;
                margin-bottom: 20px;
                border-left: 4px solid #3498db;
            }}
            .chart-description p {{ margin: 8px 0; line-height: 1.5; }}
            .chart-container {{ text-align: center; margin: 20px 0; }}
            .chart-container img {{ max-width: 100%; height: auto; border-radius: 8px; }}
            
            /* Статистическая интерпретация */
            .stat-card {{
                background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%);
                border-radius: 16px;
                padding: 20px;
                margin-top: 20px;
                border-left: 4px solid #3498db;
            }}
            .stat-card p {{ margin: 10px 0; font-size: 14px; line-height: 1.6; }}
            
            /* Формула регрессии */
            .regression-formula {{
                background: #f0f7ff;
                color: #2c3e50;
                padding: 10px 15px;
                border-radius: 10px;
                font-family: 'Courier New', monospace;
                font-size: 14px;
                text-align: center;
                margin: 10px 0;
                border: 1px solid #3498db;
            }}
            
            .correlation-value {{
                font-weight: bold;
                font-size: 16px;
                padding: 2px 8px;
                border-radius: 20px;
                display: inline-block;
            }}
            .correlation-strong {{ background: #d4edda; color: #155724; }}
            .correlation-moderate {{ background: #fff3cd; color: #856404; }}
            .correlation-weak {{ background: #f8d7da; color: #721c24; }}
            .conclusion-strong {{ color: #27ae60; font-weight: bold; }}
            .conclusion-weak {{ color: #e74c3c; font-weight: bold; }}
            
            /* Статистическая сводка */
            .summary-card {{
                background: linear-gradient(135deg, #ffffff 0%, #f8f9fa 100%);
                border-radius: 20px;
                padding: 24px;
                margin-bottom: 30px;
                border: 1px solid #e0e0e0;
            }}
            .summary-title {{
                text-align: center;
                margin-bottom: 25px;
                font-size: 20px;
                font-weight: bold;
                color: #2c3e50;
            }}
            .summary-grid {{
                display: grid;
                grid-template-columns: repeat(4, 1fr);
                gap: 20px;
                margin-bottom: 25px;
            }}
            .summary-item {{
                background: linear-gradient(135deg, #667eea15 0%, #764ba215 100%);
                border-radius: 16px;
                padding: 18px 12px;
                text-align: center;
            }}
            .summary-label {{
                font-size: 12px;
                color: #7f8c8d;
                margin-bottom: 10px;
                text-transform: uppercase;
            }}
            .summary-value {{
                font-size: 28px;
                font-weight: bold;
                font-family: 'Courier New', monospace;
                color: #2c3e50;
            }}
            .summary-interpretation {{
                background: rgba(52, 152, 219, 0.08);
                border-radius: 12px;
                padding: 18px;
                margin-top: 10px;
                line-height: 1.7;
                border-left: 3px solid #3498db;
            }}
            .highlight-number {{ color: #e67e22; font-weight: bold; }}
            .highlight-positive {{ color: #27ae60; font-weight: bold; }}
            .highlight-negative {{ color: #e74c3c; font-weight: bold; }}
            
            /* Таблица последних моделей */
            .models-table {{
                width: 100%;
                border-collapse: collapse;
                margin-top: 20px;
            }}
            .models-table th {{
                background: #f8f9fa;
                padding: 12px;
                text-align: left;
                border-bottom: 2px solid #dee2e6;
                font-weight: 600;
            }}
            .models-table td {{
                padding: 12px;
                border-bottom: 1px solid #dee2e6;
            }}
            .effect-high {{
                background: #d4edda;
                color: #155724;
                padding: 4px 12px;
                border-radius: 20px;
                font-weight: bold;
                display: inline-block;
            }}
            .effect-medium {{
                background: #fff3cd;
                color: #856404;
                padding: 4px 12px;
                border-radius: 20px;
                font-weight: bold;
                display: inline-block;
            }}
            .effect-low {{
                background: #f8d7da;
                color: #721c24;
                padding: 4px 12px;
                border-radius: 20px;
                font-weight: bold;
                display: inline-block;
            }}
            
            /* Заголовок */
            .hero-title {{
                font-size: 32px;
                font-weight: bold;
                margin-bottom: 8px;
                color: #2c3e50;
                text-align: center;
            }}
            .hero-subtitle {{
                color: #7f8c8d;
                font-size: 14px;
                text-align: center;
                margin-bottom: 30px;
            }}
            .hero-subtitle strong {{ color: #3498db; }}
            
            .date {{ text-align: right; color: #7f8c8d; font-size: 11px; margin-bottom: 20px; }}
            .footer {{ text-align: center; margin-top: 40px; padding-top: 20px; border-top: 1px solid #e0e0e0; font-size: 10px; color: #7f8c8d; }}
        </style>
    </head>
    <body>
        <div class="date">Дата отчета: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}</div>
        
        <div class="hero-title">📊 Бизнес-дашборд</div>
        <div class="hero-subtitle">
            🔬 <strong>Исследование:</strong> Влияние структурных параметров на величину <strong>общего сетевого эффекта</strong>
        </div>
        
        <!-- KPI Карточки -->
        <div class="kpi-grid">
            <div class="kpi-card">
                <div class="kpi-value">{total_models}</div>
                <div class="kpi-label">ВСЕГО МОДЕЛЕЙ</div>
                <div class="kpi-trend">📈 Готово к анализу</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-value">{avg_effect_value}</div>
                <div class="kpi-label">СРЕДНИЙ ЭФФЕКТ</div>
                <div class="kpi-trend">{'📈 Выше среднего' if avg_effect >= 0.5 else '📉 Требуется рост'}</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-value">{directed_count}</div>
                <div class="kpi-label">НАПРАВЛЕННЫХ</div>
                <div class="kpi-trend">🔄 Ориентированные связи</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-value">{weighted_count}</div>
                <div class="kpi-label">ВЗВЕШЕННЫХ</div>
                <div class="kpi-trend">📏 С весами связей</div>
            </div>
        </div>
        
        <!-- График 1: Плотность связей -->
        <div class="chart-card">
            <div class="chart-title">📐 Влияние плотности связей на сетевой эффект</div>
            <div class="chart-description">
                <p><strong>🎯 Что показывает этот график:</strong> Зависимость между количеством связей между агентами и общей эффективностью системы. <strong>Чем больше связей, тем выше сетевой эффект</strong> — это основа закона Меткалфа.</p>
                <p><strong>📈 Интерпретация точек:</strong> Каждая цветная точка — отдельная модель. Точки в верхней правой части графика показывают модели с высокой плотностью связей и высоким сетевым эффектом.</p>
            </div>
            <div class="chart-container">
                <img src="data:image/png;base64,{density_chart}" alt="График плотности связей">
            </div>
            <div class="stat-card">
                <p><strong>📊 СТАТИСТИЧЕСКАЯ ИНТЕРПРЕТАЦИЯ:</strong></p>
                <p>• <strong>Коэффициент корреляции Пирсона:</strong> 
                    <span class="correlation-value {'correlation-strong' if density_corr > 0.7 else 'correlation-moderate' if density_corr > 0.4 else 'correlation-weak'}">
                        {density_corr:.3f}
                    </span>
                    {'— <span class="correlation-strong" style="background: none; padding: 0;">сильная положительная связь</span>' if density_corr > 0.7 else '— <span class="correlation-moderate" style="background: none; padding: 0;">умеренная положительная связь</span>' if density_corr > 0.4 else '— <span class="correlation-weak" style="background: none; padding: 0;">слабая связь</span>'}
                </p>
                <p>• <strong>Уравнение регрессии:</strong></p>
                <div class="regression-formula">
                    🧮 Сетевой эффект = <span style="color: #e67e22; font-weight: bold;">{density_slope:.6f}</span> × (Количество связей) + <span style="color: #e67e22; font-weight: bold;">{density_intercept:.4f}</span>
                </div>
                <p>• <strong>Вывод:</strong> 
                {'<span class="conclusion-strong">📈 Установлена статистически значимая положительная зависимость. Рекомендуется активное наращивание связей для достижения синергетического эффекта.</span>' if density_corr > 0.5 else '<span class="conclusion-weak">📉 Выявлена слабая корреляция. Это может означать, что не количество, а качество связей играет ключевую роль.</span>'}
                </p>
            </div>
        </div>
        
        <!-- График 2: Число агентов -->
        <div class="chart-card">
            <div class="chart-title">👥 Влияние числа агентов на сетевой эффект</div>
            <div class="chart-description">
                <p><strong>🎯 Что показывает этот график:</strong> Закон Меткалфа гласит, что ценность сети пропорциональна квадрату числа участников (n²). Здесь видно, как увеличение числа агентов влияет на сетевой эффект.</p>
                <p><strong>📈 Интерпретация точек:</strong> Каждая точка — отдельная модель. Если точки образуют восходящий тренд, это подтверждает закон Меткалфа.</p>
            </div>
            <div class="chart-container">
                <img src="data:image/png;base64,{agents_chart}" alt="График числа агентов">
            </div>
            <div class="stat-card">
                <p><strong>📊 СТАТИСТИЧЕСКАЯ ИНТЕРПРЕТАЦИЯ:</strong></p>
                <p>• <strong>Коэффициент корреляции Пирсона:</strong> 
                    <span class="correlation-value {'correlation-strong' if agents_corr > 0.7 else 'correlation-moderate' if agents_corr > 0.4 else 'correlation-weak'}">
                        {agents_corr:.3f}
                    </span>
                    {'— <span class="correlation-strong" style="background: none; padding: 0;">сильная положительная связь (подтверждение закона Меткалфа)</span>' if agents_corr > 0.7 else '— <span class="correlation-moderate" style="background: none; padding: 0;">умеренная положительная связь</span>' if agents_corr > 0.4 else '— <span class="correlation-weak" style="background: none; padding: 0;">слабая связь</span>'}
                </p>
                <p>• <strong>Уравнение регрессии:</strong></p>
                <div class="regression-formula">
                    🧮 Сетевой эффект = <span style="color: #e67e22; font-weight: bold;">{agents_slope:.6f}</span> × (Количество агентов) + <span style="color: #e67e22; font-weight: bold;">{agents_intercept:.4f}</span>
                </div>
                <p>• <strong>Вывод:</strong> 
                {'<span class="conclusion-strong">📈 Подтверждается действие закона Меткалфа. Рекомендуется стратегия масштабирования сети.</span>' if agents_corr > 0.5 else '<span class="conclusion-weak">📉 Эффект масштаба не проявляется в полной мере.</span>'}
                </p>
            </div>
        </div>
        
        <!-- График 3: Веса связей -->
        {weight_chart_html}
        
        <!-- Статистическая сводка по всем моделям -->
        <div class="summary-card">
            <div class="summary-title">📊 СТАТИСТИЧЕСКАЯ СВОДКА ПО ВСЕМ МОДЕЛЯМ</div>
            <div class="summary-grid">
                <div class="summary-item">
                    <div class="summary-label">📐 СРЕДНИЙ СТРУКТУРНЫЙ ЭФФЕКТ</div>
                    <div class="summary-value">{avg_structural_value}</div>
                </div>
                <div class="summary-item">
                    <div class="summary-label">⚡ СРЕДНИЙ ФУНКЦИОНАЛЬНЫЙ ЭФФЕКТ</div>
                    <div class="summary-value">{avg_functional_value}</div>
                </div>
                <div class="summary-item">
                    <div class="summary-label">🎯 МЕДИАННЫЙ ЭФФЕКТ</div>
                    <div class="summary-value">{median_effect_value}</div>
                </div>
                <div class="summary-item">
                    <div class="summary-label">📊 СТАНДАРТНОЕ ОТКЛОНЕНИЕ</div>
                    <div class="summary-value">σ = {std_effect_value}</div>
                </div>
                <div class="summary-item">
                    <div class="summary-label">📉 МИНИМАЛЬНЫЙ ЭФФЕКТ</div>
                    <div class="summary-value">{min_effect_value}</div>
                </div>
                <div class="summary-item">
                    <div class="summary-label">📈 МАКСИМАЛЬНЫЙ ЭФФЕКТ</div>
                    <div class="summary-value">{max_effect_value}</div>
                </div>
                <div class="summary-item">
                    <div class="summary-label">📌 25-й ПЕРЦЕНТИЛЬ</div>
                    <div class="summary-value">Q1 = {percentile_25:.4f}</div>
                </div>
                <div class="summary-item">
                    <div class="summary-label">📌 75-й ПЕРЦЕНТИЛЬ</div>
                    <div class="summary-value">Q3 = {percentile_75:.4f}</div>
                </div>
            </div>
            <div class="summary-interpretation">
                <strong>💡 ДЕТАЛЬНАЯ ИНТЕРПРЕТАЦИЯ:</strong><br><br>
                • <strong>Общий сетевой эффект</strong> в среднем составляет <span class="highlight-number">{avg_effect_value}</span> 
                {'— <span class="highlight-positive">высокий уровень</span>' if avg_effect >= 60 else '— <span class="highlight-number">средний уровень</span>' if avg_effect >= 30 else '— <span class="highlight-negative">низкий уровень</span>'}<br><br>
                • <strong>Разброс значений</strong> (σ = <span class="highlight-number">{std_effect_value}</span>) показывает, что эффективность 
                {'сильно варьируется в зависимости от структуры сети' if std_effect > 0.2 else 'относительно стабильна'}<br><br>
                • <strong>50% моделей</strong> имеют сетевой эффект в диапазоне от <span class="highlight-number">{percentile_25:.4f}</span> до <span class="highlight-number">{percentile_75:.4f}</span><br><br>
                • <strong>Потенциал роста</strong> составляет <span class="highlight-positive">{growth_potential:.4f}</span> (или <span class="highlight-number">{growth_percent:.0f}%</span> относительного улучшения)
            </div>
        </div>
        
        <div class="footer">
            <p>© 2026 NetEffect Modeler. Все права защищены.</p>
        </div>
    </body>
    </html>
    """
    
    response = HttpResponse(html_content, content_type='text/html')
    response['Content-Disposition'] = 'attachment; filename="dashboard_report.html"'
    return response